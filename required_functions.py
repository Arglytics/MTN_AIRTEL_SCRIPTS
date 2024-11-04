import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from io import StringIO
import datetime
import pyodbc
print("|-connecting to DASH20")
conn_str = (
r'DRIVER={SQL Server};'
r'SERVER=10.128.0.170;'
r'DATABASE=DASH20;'
r'UID=sa;'
r'PWD=mis2017;')
connection =pyodbc.connect(conn_str)

# MTN STATEMENT
mtn_extract=None      
mtn_files=[]
path1=None
for file in os.listdir(r"C:\Users\biugateway\OneDrive - DFCU Bank Uganda\RECON\MTN\\"):
    mtn_files.append(file)
if len(mtn_files)>0:
    if datetime.datetime.fromtimestamp(os.stat(os.path.join(r'C:\Users\biugateway\OneDrive - DFCU Bank Uganda\RECON\MTN',mtn_files[-1])).st_mtime).date().strftime("%Y-%m-%d")!=(datetime.datetime.today()).strftime("%Y-%m-%d"):
        directory=os.listdir(r"C:\Users\biugateway\OneDrive - DFCU Bank Uganda\STATEMENT\MTN\\") 
        for file in directory:
            if datetime.datetime.fromtimestamp(os.stat(os.path.join(r'C:\Users\biugateway\OneDrive - DFCU Bank Uganda\STATEMENT\MTN',file)).st_mtime).date().strftime("%Y-%m-%d")==(datetime.datetime.today()).strftime("%Y-%m-%d") and not file.lower().__contains__('image'):
                path1=os.path.join(r'C:\Users\biugateway\OneDrive - DFCU Bank Uganda\STATEMENT\MTN',file)
                mtn_extract=file 
 
# ARTEL STATEMENT
airtel_extract=None 
airtel_files=[]
path2=None
for file in os.listdir(r"C:\Users\biugateway\OneDrive - DFCU Bank Uganda\RECON\AIRTEL\\"):
    airtel_files.append(file)
if datetime.datetime.fromtimestamp(os.stat(os.path.join(r'C:\Users\biugateway\OneDrive - DFCU Bank Uganda\RECON\AIRTEL',airtel_files[-1])).st_mtime).date().strftime("%Y-%m-%d")!=(datetime.datetime.today()).strftime("%Y-%m-%d"):
    directory=os.listdir(r"C:\Users\biugateway\OneDrive - DFCU Bank Uganda\STATEMENT\AIRTEL\\") 
    for file in directory:
        if datetime.datetime.fromtimestamp(os.stat(os.path.join(r'C:\Users\biugateway\OneDrive - DFCU Bank Uganda\STATEMENT\AIRTEL',file)).st_mtime).date().strftime("%Y-%m-%d")==(datetime.datetime.today()).strftime("%Y-%m-%d") and not file.lower().__contains__('image'):
            path2=os.path.join(r'C:\Users\biugateway\OneDrive - DFCU Bank Uganda\STATEMENT\AIRTEL',file)
            airtel_extract=file


try:
    class Recon:
        def __init__(self):
            print("STARTING RECON")
        def mtn_recon(self,filepath,filename):
            """
            FUNCTION TO READ AND RECONCILE AIRTEL STATEMENT AND LEDGER
            """
            
            if filename is not None and  len(str(filename))>0 and filepath is not None:
                print('|'+"="*20+"READING MTN STATEMENT")
                list=[]
                with open(filepath) as file_:
                    data=file_.readlines()[3:]
                    for lines in data:
                        for line in lines.split('\n'):
                            list.append(line)
                #putting all elements inthe lists together 
                statement='\n'.join(list)
                statement=pd.read_csv(StringIO(statement))
                #LEDGER
                print('|'+"="*20+"READING FINANCLE LEDGER")           
                ledger="""
                SELECT CAST(TRANSACTION_DATE AS DATE)"DATE"
                ,TRANSACTION_ID "TRAN_ID"
                ,NARRATIVE "NARRATION"
                ,TRANSACTION_AMOUNT_BASE "AMOUNT" 
                FROM TRANSACTIONS 
                WHERE TRANSACTION_DATE=CAST(GETDATE()-1 AS DATE) AND ACCOUNT_NUMBER='01001106100012'
                """
                ledger=pd.read_sql(ledger,connection)
                #MAIPULATING LEDGER
                print('|'+"="*20+"CLEANING FINANCLE LEDGER")  
                ledger['TELCO_ID']=ledger['NARRATION'].str.rsplit('-').apply(lambda x: x[1] if len(x)==3 else '')
                ledger['TELCO_ID']=ledger['TELCO_ID'].replace('',method='bfill')
                ledger.rename(columns={'AMOUNT':'LEDGER_AMOUNT'},inplace=True)
                summed_amount=ledger.groupby(['TELCO_ID'])['LEDGER_AMOUNT'].sum() #SUMMING TO GET THE TOTALS
                recon=pd.DataFrame(summed_amount).reset_index()
                #MAIPULATING STATEMENT
                print('|'+"="*20+"CLEANING MTN STATEMENT")  
                statement_to_use=statement[['DATE_','FID','EXTID','FROMMSISDN','TRANSTYPE','AMOUNT']].copy()
                statement_to_use.FID=statement_to_use.FID.map(str)
                statement_to_use.rename(columns={"AMOUNT":"STATEMENT_AMOUNT"},inplace=True)
                #MERGING STATEMENT AND LEDGER 
                print('|'+"="*20+"MERGING STATEMENT AND LEDGER FOR RECON") 
                recon=statement_to_use.merge(recon, how='left',left_on='FID',right_on='TELCO_ID')
                recon['DIFF']=recon['STATEMENT_AMOUNT']-recon['LEDGER_AMOUNT']
                #CAPTURING EXCEPTIONS
                print('|'+"="*20+"CAPTURING EXCEPTIONS") 
                check1=recon.query("DIFF>0") # checking for exception 1 difference less greater than 0
                #getting outstanding in the ledger for check 1
                ledgeroutstanding_check1=ledger[ledger.TELCO_ID.isin(check1.TELCO_ID.to_list())]
                #matched in the ledger for check 1
                ledgermatched_check1=ledger[~ledger.TELCO_ID.isin(check1.TELCO_ID.to_list())]
                #getting outstanding in statement for check1
                statementoutstanding_check1=statement_to_use[statement_to_use.FID.isin(check1.TELCO_ID.to_list())]
                #getting matched in statatement for check 1
                statementmatched_check1=statement_to_use[~statement_to_use.FID.isin(check1.TELCO_ID.to_list())]
                #checking for exception 2 in statement ,those with trantype credit and amount >0
                #getting oustanding in statement for check 2
                statementoutstanding_check2=statementmatched_check1[statementmatched_check1['TRANSTYPE']=='CREDIT'].query("STATEMENT_AMOUNT>0")
                #getting matched for check 2
                statementmatched_check2=statementmatched_check1[~statementmatched_check1.FID.isin(statementoutstanding_check2.FID.to_list())]
                #checking for exception 3 in statement ,those with negative amount
                statementoutstanding_check3=statementmatched_check2[statementmatched_check2.STATEMENT_AMOUNT.map(str).str.contains('-')]
                #getting the matched for check 3
                statementmatched_check3=statementmatched_check2[~statementmatched_check2.FID.isin(statementoutstanding_check3.FID.to_list())]
                #checking for all the fids instatement not in telco for ledger
                statementoutstanding_check4=statementmatched_check3[~statementmatched_check3['FID'].isin(ledgermatched_check1.TELCO_ID.to_list())]
                #getting final matched for MTN
                statementmatched_check4=statementmatched_check3[statementmatched_check3['FID'].isin(ledgermatched_check1.TELCO_ID.to_list())]
                #getting check 2 for ledger
                ledgeroutstanding2=ledgermatched_check1[~ledgermatched_check1.TELCO_ID.isin(statementmatched_check4.FID.to_list())]
                ledgermatched_check2=ledgermatched_check1[ledgermatched_check1.TELCO_ID.isin(statementmatched_check4.FID.to_list())]
                #last checks for ledger 
                ledgeroutstanding3=ledgermatched_check2[ledgermatched_check2.NARRATION.map(str).str.lower().str.contains('rtgs')]
                ledgermatched3=ledgermatched_check2[~ledgermatched_check2.TRAN_ID.isin(ledgeroutstanding3.TRAN_ID.to_list())].drop("TELCO_ID",axis=1)
                #merging outstanding
                statement_outstanding=pd.concat([statementoutstanding_check1,statementoutstanding_check2,statementoutstanding_check3,statementoutstanding_check4]).reset_index(drop=True).rename(columns={'AMOUNT':"STATEMENT_AMOUNT"})
                # Ledger_statement_outstanding=statement_outstanding.merge(ledgeroutstanding_check1,how="left",left_on='FID',right_on='TELCO_ID').rename(columns={'AMOUNT':"LEDGER_AMOUNT"})
                # Ledger_statement_outstanding.insert(6,'__','')
                print('|'+"="*20+"EXTRACTING RECON FILE") 
                #EXTARCTING FILE
                date=ledgeroutstanding_check1['DATE'].dt.date.to_list()[0]
                #creating the excel file
                # os.makedirs(r'C:\Users\RSenyonjo\Downloads\MTN_UNI_RECONS')
                wb=openpyxl.Workbook()
                excel_filename=r'C:\Users\biugateway\OneDrive - DFCU Bank Uganda\RECON\MTN\MTN_UNI_RECON_{}.xlsx'.format(date)
                wb.save(excel_filename)
                wb=openpyxl.load_workbook(excel_filename)
                wb.create_sheet('MATCHED_LEDGER')
                wb.create_sheet('STATEMENT_OUTSTANDING')
                wb.create_sheet('LEDGER_OUTSTANDING')
                wb.save(excel_filename)
                ws1=wb['STATEMENT_OUTSTANDING']
                wb['Sheet'].title='MATCHED_STATEMENT'
                ws2=wb['MATCHED_STATEMENT']
                ws3=wb['LEDGER_OUTSTANDING']
                ws4=wb['MATCHED_LEDGER']
                if statementmatched_check4.STATEMENT_AMOUNT.sum()-ledgermatched_check1.LEDGER_AMOUNT.sum()==0:
                    for row in dataframe_to_rows(statement_outstanding,index=False,header=True):
                        ws1.append(row)
                    for row in dataframe_to_rows(statementmatched_check4,index=False,header=True):
                        ws2.append(row)
                    for row in dataframe_to_rows(pd.concat([ledgeroutstanding_check1,ledgeroutstanding2,ledgeroutstanding3]),index=False,header=True):
                        ws3.append(row)
                    for row in dataframe_to_rows(ledgermatched3,index=False,header=True):
                        ws4.append(row)
                    wb.save(excel_filename)
                    print("MTN_RECON FOR {}  EXTRACTED".format(date))
                else:
                    return "+"*5+"MTN STATEMENT LEDGER NOT PROPERLY RECONCILED"+"+"*5+'\n\n'+'='*50
            else:
                print("+"*4+" NO MTN STATEMENT SENT TODAY OR FILE WAS RECONCILED "+"+"*4)
                

        def airtel_recon(self,filepath,filename):
            """
            FUNCTION TO READ AND RECONCILE AIRTEL STATEMENT AND LEDGER
            """
            if filename is not None and len(str(filename))>0 and filepath is not None:
                print('|'+"="*20+"READING AIRTEL STATEMENT")
                astatement=pd.read_excel(path2)
                astatement.columns=astatement.columns.str.replace(' ','_')
                astatement=astatement[['Transaction_ID', 'External_Reference', 'Transaction_Date','Sender_Mobile_Number','Payer_Details','Service_Type','Status','Transaction_Amount']].copy()
                astatement['Transaction_ID']=astatement['Transaction_ID'].map(str)
                print('|'+"="*20+"READING AIRTEL LEDGER")
                ledger="""
                SELECT CAST(TRANSACTION_DATE AS DATE)"DATE"
                ,TRANSACTION_ID "TRAN_ID"
                ,NARRATIVE "NARRATION"
                ,TRANSACTION_AMOUNT_BASE "LEDGER_AMOUNT" 
                FROM TRANSACTIONS 
                WHERE TRANSACTION_DATE=CAST(GETDATE()-1 AS DATE) AND ACCOUNT_NUMBER='01001160100013'
                """
                airtel_ledger=pd.read_sql(ledger,connection)
                print('|'+"="*20+"CLEANING AIRTEL LEDGER")
                airtel_ledger['TELCO']=airtel_ledger['NARRATION'].str.rsplit('-').apply(lambda x : x[1] if len(x)==3 else '')
                airtel_ledger.TELCO.replace('',method='bfill',inplace=True)
                ledger_summed=airtel_ledger.groupby('TELCO')['LEDGER_AMOUNT'].sum().reset_index()
                print('|'+"="*20+"MERGING AIRTEL STATEMENT AND LEDGER FOR RECON")
                recon=astatement.merge(ledger_summed, how='left',left_on='Transaction_ID',right_on='TELCO')
                recon['DIFF']=recon.Transaction_Amount-recon.LEDGER_AMOUNT
                print('|'+"="*20+'capturing the exceptions'.upper())
                # check 1 where difference >0
                check1=recon[recon.DIFF>0]
                #getting the outstanding the leger and statatement 
                #ledger
                ledger_outstanding1=airtel_ledger[airtel_ledger['TELCO'].isin(check1.Transaction_ID.to_list())]
                #statement
                astatement_outstanding1=astatement[astatement['Transaction_ID'].isin(check1.Transaction_ID.to_list())]
                #getting the matched legder and statement 
                ledger_matched1=airtel_ledger[~airtel_ledger['TELCO'].isin(check1.Transaction_ID.to_list())]
                astatement_matched1=astatement[~astatement['Transaction_ID'].isin(check1.Transaction_ID.to_list())]
                # check2 where service type not merchant type and status= transaction success 
                astatement_outstanding2=astatement_matched1[(astatement_matched1['Service_Type']!='Merchant Payment') & (astatement_matched1['Status']=='Transaction Success')]
                #matched statement for check 2
                astatement_matched2=astatement_matched1[~astatement_matched1.Transaction_ID.isin(astatement_outstanding2.Transaction_ID.to_list())]
                #check 3 where status not transaction success
                astatement_outstanding3=astatement_matched2[astatement_matched2['Status']!='Transaction Success']
                #matched statement for check 3
                astatement_matched3=astatement_matched2[~astatement_matched2.Transaction_ID.isin(astatement_outstanding3.Transaction_ID.to_list())]
                #check 4 tid in statement but not in ledger
                astatement_outstanding4=astatement_matched3[~astatement_matched3['Transaction_ID'].isin(ledger_matched1['TELCO'].to_list())]
                #matched statement for 4 
                astatement_matched4=astatement_matched3[~astatement_matched3.Transaction_ID.isin(astatement_outstanding4.Transaction_ID.to_list())]
                #check 5 tid in ledger but not in statement
                ledger_outstanding2=ledger_matched1[~ledger_matched1['TELCO'].isin(astatement_matched4.Transaction_ID.to_list())]
                ledger_matched2=ledger_matched1[~ledger_matched1.TELCO.isin(ledger_outstanding2.TELCO.to_list())]
                print('|'+"="*20+"EXTRACTING THE FILE")
                date=pd.to_datetime(astatement.Transaction_Date).dt.date.to_list()[0]
                wb=openpyxl.Workbook()
                excel_file=r'C:\Users\biugateway\OneDrive - DFCU Bank Uganda\RECON\AIRTEL\AIRTEL_UNI_RECON_{}.xlsx'.format(date)
                wb.save(excel_file)
                wb=openpyxl.load_workbook(excel_file)
                wb[wb.sheetnames[0]].title='AIRTEL_STATEMENT_MATCHED'
                wb.create_sheet('FINACLE_AIRTEL_MATCHED')
                wb.create_sheet('STATEMENT_OUTSTANDING')
                wb.create_sheet('LEDGER_OUTSTANDING')
                wb.save(excel_file)
                as1=wb['AIRTEL_STATEMENT_MATCHED']
                as2=wb['FINACLE_AIRTEL_MATCHED']
                as3=wb['STATEMENT_OUTSTANDING']
                as4=wb['LEDGER_OUTSTANDING']
                if ledger_matched2['LEDGER_AMOUNT'].sum()-astatement_matched4.Transaction_Amount.sum() == 0:
                    for row in dataframe_to_rows(astatement_matched4,index=False,header=True):
                        as1.append(row)
                    for row in dataframe_to_rows(ledger_matched2,index=False,header=True):
                        as2.append(row)
                    for row in dataframe_to_rows(pd.concat([astatement_outstanding1,astatement_outstanding2,astatement_outstanding3,astatement_outstanding4]),index=False,header=True):
                        as3.append(row)
                    for row in dataframe_to_rows(pd.concat([ledger_outstanding1,ledger_outstanding2]),index=False,header=True):
                        as4.append(row)
                    wb.save(excel_file)
                    print("AIRTEL_RECON FOR {}  EXTRACTED".format(date))
                else:
                    print("+"*5+"AIRTEL STATEMENT LEDGER NOT PROPERLY RECONCILED"+"+"*5)
            else:
                print("+"*4+" NO AIRTEL STATEMENT FOR YESTERDAY OR FILE WAS RECONCILED "+"+"*4)
except:
    print("+"*5+"failure"+"+"*5)



#EXECUTING THE FUNCTIONS
recon_=Recon()
recon_.mtn_recon(filepath=path1,filename=mtn_extract)
recon_.airtel_recon(filepath=path2,filename=airtel_extract)
